import dateutil.parser
import random
import string
import json

from django.contrib.auth.models import AnonymousUser
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from pytz import timezone
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string

from rest_framework import views, status
from rest_framework.decorators import list_route, detail_route
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response

from core import viewsets
from payment.models import Payment
from organization.models import AccessRequest
from .models import User, Group, Note, Diploma
from .serializers import UserSerializer, UserWriteSerializer, GroupSerializer, NoteWriteSerializer, NoteSerializer, \
    DiplomaSerializer, DiplomaWriteSerializer


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = []

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return UserSerializer
        else:
            return UserWriteSerializer

    def get_queryset(self):
        qs = User.on_site.filter(is_active=True)
        q = Q()
        # Преподаватель видит только своих учеников
        if self.request.user and self.request.user.__class__ != AnonymousUser and self.request.user.role == 'teacher':
            q &= Q(role='teacher') | Q(role='admin') | (Q(role='student') & Q(groups__author=self.request.user))
        data = self.request.query_params

        if data.get('role'):
            if data.get('role') == 'teacher':
                q &= Q(role='teacher')  # | Q(role='admin')
            else:
                q &= Q(role=data.get('role'))
        
        if self.request.GET.get('filter'):
            scope_filters = json.loads(self.request.GET.get('filter'))['role']            
            q &= Q(role__in=scope_filters)           
        
              
        if self.request.GET.get('page'):  
            limit = 20
            offset = int(self.request.GET.get('page'))
            #last_id -= limit * offset
            qs = qs.filter(q).order_by('-id')[offset*limit:(offset+1)*limit]
        else:        
            qs = qs.filter(q).order_by('-id')       
            
        return qs

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        user.is_active = False
        user.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @detail_route(methods=['POST'])
    def password_change(self, request, pk=None):
        user = User.objects.get(id=int(pk))
        if not user:
            return Response({'error': 'Пользователь с данным id не найден'})

        user.set_password(request._data['password'])
        user.save()
        return Response(status=status.HTTP_200_OK)

    @detail_route(methods=['GET'])
    def my_student(self, request, pk=None):
        if (request.user.__class__ is AnonymousUser) or (request.user.role == 'admin'):
            result = {'result': True}
        else:
            result = {'result': AccessRequest.objects.filter(user__id=int(pk), course__authors=request.user).exists()}
        return Response(result)

    @list_route(methods=['GET', 'POST'])
    def activate(self, request, pk=None):
        try:
            code = request._request.environ.get('QUERY_STRING', None).split('=')[1]
            user = User.objects.get(unsubscribe_code=code)
            user.is_active = True
            user.save()
            return Response(status=status.HTTP_200_OK)
        except:
            return Response(status=status.HTTP_404_NOT_FOUND)

    @list_route(methods=['GET'])
    def profile(self, request):
        if request.user.is_authenticated():
            serializer = self.serializer_class(request.user)
            return Response(serializer.data)

        else:
            return Response(
                {"role": {"value": "anonymous", "title": "anonymous"}, "avatar": "/static/images/default-profile.jpg"})

    @list_route(methods=['GET'])
    def export(self, request, pk=None):
        ids = list(map(int, request.query_params.getlist('users')))
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="export.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Пользователи'
        ws.column_dimensions['A'].width = 50.0
        ws.column_dimensions['B'].width = 50.0
        ws.column_dimensions['C'].width = 50.0
        ws.column_dimensions['D'].width = 50.0

        header = ['ФИО', 'Роль', 'Дата регистрации', 'Email']
        ws.append(header)

        for user in User.objects.filter(site=request.site, is_active=True, id__in=ids):
            ws.append([user.full_name,
                       user.get_role_display(),
                       user.registered_at.strftime('%d.%m.%Y'),
                       user.email])

        wb.save(response)
        return response

    @list_route(methods=['POST'])
    def register(self, request, format='json'):
        email = request.data.get('email').lower()

        if not User.objects.filter(site=request.site, email__iexact=email).exists():
            password = request.data.get('password')
            if password is None:
                password = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(8))

            user = User.objects.create_user(site=request.site,
                                            is_active=True,
                                            role=request.data.get('role'),
                                            email=email,
                                            password=password,
                                            first_name=request.data.get('first_name', ''),
                                            middle_name=request.data.get('middle_name', ''),
                                            last_name=request.data.get('last_name', ''),
                                            city=request.data.get('city', ''),
                                            grade=request.data.get('grade', ''),
                                            speciality=request.data.get('speciality', ''),
                                            gender=request.data.get('gender', ''),
                                            examination=request.data.get('examination', ''),
                                            phone=request.data.get('phone', ''))

            # Values for custom fields are stored in CustomField model in 'values' field, so
            # we need to save it in another way:
            for field in request.data.get('custom_fields', []):
                # [] is for AddUserModal (multiple users)
                # request.custom_fields is a list of [name, value] for each field
                custom_field = user.site.organization.custom_fields.get(name=field[0])
                value = field[1]
                if custom_field.field_type == 6:  # 6 is DateField
                    # Currently we get time in UTC from frontend, so we need to represent it in local TZ
                    # To store time in each user's local timezone, change timezone() below
                    value = dateutil.parser.parse(value).astimezone(timezone('Europe/Moscow')).date()
                elif custom_field.field_type == 7:  # 7 is DateTimeField
                    value = dateutil.parser.parse(value).astimezone(timezone('Europe/Moscow'))
                custom_field.values.update({user.email: str(value)})
                custom_field.save()

            if request.data.get('groups'):
                user.groups.add(*request.data.get('groups'))

            if request.data.get('tags'):
                user.tags.add(request.data.get('tags'))

            # Email for new user
            params = {
                'address': request.site.domain + '.grandclass.net',
                'user': user,
                'password': password
            }

            if user.role == 'student':
                email_body = render_to_string('mail/new_student_registered.txt', params)
            else:
                email_body = render_to_string('mail/new_teacher_registered.txt', params)

            user.email_user(request.site.organization.title, email_body, settings.DEFAULT_FROM_EMAIL)

            # Email for admin
            params = {'user': user}
            admin = request.site.organization.admin
            if admin and request.site.organization.notify_about_clients:
                if user.role == 'student':
                    email_body = render_to_string('mail/new_user_notification_for_platform_admin.txt', params)
                    admin.email_user('Новый ученик', email_body, settings.DEFAULT_FROM_EMAIL)
                else:
                    email_body = render_to_string('mail/new_teacher_notification_for_platform_admin.txt', params)
                    admin.email_user('Новый преподаватель', email_body, settings.DEFAULT_FROM_EMAIL)

            return Response(status=status.HTTP_201_CREATED, data={'id': user.id})
        else:
            return Response({'status': 209})

    @list_route(methods=['POST'])
    def login(self, request, format=None):
        data = request.data

        email = data.get('email', '').lower()
        password = data.get('password', '')
        user = authenticate(username=email, password=password)

        if User.objects.filter(email=email).first():
            if not User.objects.filter(email=email).first().is_active:
                return Response({'status': status.HTTP_403_FORBIDDEN})

        if user is not None:
            login(request, user)
            return Response({'status': status.HTTP_200_OK})
        else:
            return Response({'status': status.HTTP_404_NOT_FOUND})

    @list_route(methods=['POST'])
    def password_reset(self, request):
        email = request.data.get('email', None)

        if email is None:
            return Response({'status': status.HTTP_404_NOT_FOUND})

        if User.objects.filter(site=request.site, email__iexact=email).exists():
            user = User.objects.get(site=request.site, email__iexact=email)
            random_password = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(8))
            user.set_password(random_password)
            user.save()

            user.email_user(subject='Сброс пароля Системы электронного обучения ГБУ ГППЦ ДОгМ',
                            message='Вы получили это письмо, потому что Вы (либо кто-то, выдающий себя за Вас) отправил запрос на смену пароля для доступа к Личному кабинету в системе электронного обучения ГБУ ГППЦ ДОгМ.\n\nЕсли Вы не отправляли подобный запрос, то не обращайте внимания на это письмо или обратитесь в службу поддержки.\nВаш новый пароль: {}\nСсылка для входа http://sdo.gppc.ru/login\n\n\nС уважением,\nСлужба поддержки пользователей\nСистемы электронного обучения ГБУ ГППЦ ДОгМ\nАдрес: Есенинский бульвар, дом 12, корпус 2\nИнтернет представительство: http://sdo.gppc.ru/\nЭлектронная почта: sdo@gppc.ru'.format(
                                random_password),
                            from_email=settings.DEFAULT_FROM_EMAIL)

            return Response({'status': status.HTTP_200_OK})
        else:
            return Response({'status': status.HTTP_404_NOT_FOUND})

    @list_route(methods=['POST'])
    def batch_reset(self, request):
        ids = []
        query = self.request.data['ids'].split('&')[:-1]
        for each in query:
            ids.append(each.split('=')[1])

        for user in User.objects.filter(id__in=ids):
            random_password = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(8))
            user.set_password(random_password)
            user.save()

            user.email_user(subject='Новые данные для входа',
                            message='Ваш пароль был сброшен, новый пароль: {}\nСсылка для входа http://sdo.gppc.ru/login \nПароль можно сменить в настройках профиля.'.format(random_password),
                            from_email=settings.DEFAULT_FROM_EMAIL)
        return Response(status=status.HTTP_200_OK)

    @list_route(methods=['POST'])
    def batch_delete(self, request):
        queryset = self.get_queryset()
        data = self.request.data
        queryset.filter(id__in=data['ids']).update(is_active=False)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @list_route(methods=['GET'])
    def sales(self, request):
        try:
            query = request._request.environ.get('QUERY_STRING', None).split('&')
            params = [param.split('=')[1] for param in query]
            date_end = datetime.combine(datetime.strptime(params[0], '%d.%m.%Y').date(),
                                        datetime.min.time()) + timedelta(days=1)
            date_start = datetime.combine(datetime.strptime(params[1], '%d.%m.%Y').date(), datetime.min.time())
        except:
            return Response()

        result = []
        for teacher in User.objects.filter(role='teacher'):
            data = {
                'id': teacher.id,
                'full_name': teacher.full_name,
                'courses': 0,
                'webinars': 0
            }

            for payment in Payment.objects.filter(is_paid=True, paid_at__gte=date_start, paid_at__lte=date_end):
                if teacher in payment.access_requests.last().target.authors.all():
                    if payment.access_requests.last().target.__class__.__name__ == 'Course':
                        data['courses'] += payment.amount
                    if payment.access_requests.last().target.__class__.__name__ == 'Webinar':
                        data['webinars'] += payment.amount

            if data['courses'] or data['webinars']:
                result.append(data)
        return Response(result)


class UsersImportView(views.APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, filename, format=None):
        try:
            file_obj = request.FILES.get('file')
            workbook = load_workbook(file_obj)
            first_sheet = workbook.get_sheet_names()[0]
            worksheet = workbook.get_sheet_by_name(first_sheet)
        except:
            return Response(status=status.HTTP_201_CREATED, data={'error': 'Загружен некорректный файл', 'counts': {}})

        for row in list(worksheet.rows)[1:]:
            if not (row[0].value and row[2].value and row[3].value):
                return Response(status=status.HTTP_201_CREATED, data={'error': 'Поля имя, фамилия и email являются обязательными', 'counts': {}})

        count_created = 0
        count_failed = 0

        for row in list(worksheet.rows)[1:]:
            try:
                first_name = row[0].value
                middle_name = row[1].value
                last_name = row[2].value
                email = row[3].value
                phone = row[4].value

                if not User.objects.filter(email__iexact=email, site=request.user.site).exists():
                    random_password = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(8))

                    user = User.objects.create_user(site=request.user.site,
                                                    email=email,
                                                    password=random_password,
                                                    first_name=first_name,
                                                    middle_name=middle_name if middle_name else '',
                                                    last_name=last_name,
                                                    phone=phone if phone else '',
                                                    role='student')
                    count_created += 1

                    try:
                        # Email for new user
                        params = {
                            'address': request.site.domain + '.grandclass.net',
                            'user': user,
                            'password': random_password
                        }

                        email_body = render_to_string('mail/new_student_registered.txt', params)
                        user.email_user(request.site.organization.title, email_body, settings.DEFAULT_FROM_EMAIL)

                        # Email for admin
                        admin = request.site.organization.admin
                        if admin and request.site.organization.notify_about_clients:
                            email_body = render_to_string('mail/new_user_notification_for_platform_admin.txt', {'user': user})
                            admin.email_user('Новый ученик', email_body, settings.DEFAULT_FROM_EMAIL)
                    except:
                        pass

            except:
                count_failed += 1

        return Response(status=status.HTTP_201_CREATED, data={'error': '', 'counts': {'created': count_created, 'failed': count_failed}})


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = []

    def get_queryset(self):
        qs = Group.on_site.all()
        data = self.request.query_params

        q = Q()
        if data.get('course'):
            try:
                q &= Q(course=int(data.get('course')))
            except ValueError:
                pass
        qs = qs.filter(q)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save(author=self.request.user)
        instance.users.add(*self.request.data.get('users', None))

    def perform_update(self, serializer):
        instance = serializer.save()
        old_users = set([user.id for user in instance.users.all()])

        try:
            new_users = set(self.request.data.get('users', None))
        except:
            new_users = set()

        instance.users.add(*new_users - old_users)
        instance.users.remove(*old_users - new_users)

    @list_route(methods=['POST'])
    def batch_delete(self, request):
        queryset = self.get_queryset()
        data = self.request.data
        queryset.filter(id__in=data['ids']).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class NoteViewSet(viewsets.ModelViewSet):
    queryset = Note.objects.all()
    serializer_class = NoteSerializer
    permission_classes = []

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return NoteSerializer
        return NoteWriteSerializer

    def perform_create(self, serializer):
        serializer.save(author=self.request.user)

    def get_queryset(self):
        qs = Note.on_site.all()
        user = self.request.user
        data = self.request.query_params
        if self.request.user.__class__ is AnonymousUser:
            role = 'student'
        else:
            role = user.role
        if role == 'student':
            qs = Note.on_site.none()

        q = Q()
        if data.get('user'):
            q &= Q(user__id=data.get('user'))
        qs = qs.filter(q).order_by('created_at')

        if role == 'teacher':
            if data.get('user'):
                check = False
                student = User.objects.get(id=data.get('user'))
                for request in student.access_requests.filter(course__isnull=False):
                    if user in request.course.authors.all():
                        check = True

                for group in student.groups.all():
                    for request in group.access_requests.filter(course__isnull=False):
                        if user in request.course.authors.all():
                            check = True

                if not check:
                    qs = Note.on_site.none()

        return qs


class DiplomaViewSet(viewsets.ModelViewSet):
    queryset = Diploma.objects.all()
    serializer_class = DiplomaSerializer
    permission_classes = []

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return DiplomaSerializer
        return DiplomaWriteSerializer

    def get_queryset(self):
        qs = Diploma.on_site.all()
        user = self.request.user
        data = self.request.query_params

        q = Q()
        if 'my' in data:
            q &= Q(user=user)

        if 'user' in data:
            try:
                q &= Q(user__id=int(data.get('user')))
            except ValueError:
                pass

        qs = qs.filter(q).order_by('id')
        return qs
